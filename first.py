import openpyxl
from openpyxl.styles.builtins import total

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
product_under_10 = {}
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    price = product_list.cell(product_row, 3).value
    inventory = product_list.cell(product_row, 2).value
    product_num = product_list.cell(product_row, 1).value

    # calculation for number of products per supplier
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] = products_per_supplier[supplier_name] + 1
    else:
        products_per_supplier[supplier_name] = 1

    # calculation for total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
       total_value_per_supplier[supplier_name] = total_value_per_supplier[supplier_name] + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    #logic to find products with inventory less than 10
    if inventory < 10 :
        product_under_10[int(product_num)] = int(inventory)

    #add value for total inventory price
    product_list.cell(product_row, 5).value = price * inventory

inv_file.save("new_inventory.xlsx")